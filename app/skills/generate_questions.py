import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.config import settings
from app.services.storage_service import storage_service
from app.services.task_service import task_service
from app.services.question_service import question_service
from app.skills.task_update_status import task_update_status
from app.utils.logger import setup_logger

# ===== 把下面 3 个 import 改成你实际放置的位置 =====
from app.utils.cut_questions_by_ocr import cut_questions_by_tencent_ocr
from app.utils.cut_solutions import extract_analysis_by_question_blocks
from app.utils.clean_analysis_by_ocr import clean_analysis_image

logger = setup_logger(settings.log_level, settings.logs_dir)


def _sorted_png_files(folder: Path) -> list[Path]:
    return sorted([p for p in folder.glob("*.png") if p.is_file()])


def _clear_png_files(folder: Path) -> None:
    for p in folder.glob("*.png"):
        if p.is_file():
            p.unlink()


def _clear_dir(folder: Path) -> None:
    if folder.exists():
        shutil.rmtree(folder)
    folder.mkdir(parents=True, exist_ok=True)


def _sanitize_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", "", name)
    return name or "question"


def _read_excel_question_labels(excel_path: Path) -> list[str]:
    """
    从 Excel 中读取题号标签。
    这里按你的现有表格习惯，优先读取第一列中像“填空题1 / 计算题2 / 解决问题7”这样的标签。
    如果未来表结构变了，再单独改这个函数即可。
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    labels: list[str] = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        first_cell = row[0]
        if first_cell is None:
            continue

        text = str(first_cell).strip()
        if not text:
            continue

        # 典型标签：填空题1 / 计算题2 / 选择题3 / 解决问题7
        if re.search(r"(题|问题)\s*\d+$", text) or re.search(r"\d+$", text):
            # 排除明显不是题号的头部元数据
            if text in {"学校", "试题时间", "试卷说明", "题号"}:
                continue
            labels.append(text)

    if not labels:
        raise ValueError(f"Excel 中未读取到题号标签: {excel_path}")

    return labels


def _collect_cut_results(page_output_dir: Path, prefix_pattern: str) -> list[Path]:
    """
    从单页切割目录中收集结果图，按 question_1 / question_2 排序。
    """
    files = [p for p in page_output_dir.glob("*.png") if p.is_file()]
    valid_files = []

    for p in files:
        if p.name == "debug_lines.png":
            continue
        if p.name.endswith("_clean.png"):
            continue
        if re.match(prefix_pattern, p.name):
            valid_files.append(p)

    def sort_key(path: Path):
        m = re.search(r"question_(\d+)", path.stem)
        return int(m.group(1)) if m else 10**9

    return sorted(valid_files, key=sort_key)


def generate_questions(task_id: str, excel_path: str) -> dict[str, Any]:
    """
    Step10.4 总调度：
    1. blank_pages -> blank_questions
    2. solution_pages -> solution_questions
    3. 清洗 solution_questions（直接覆盖）
    4. Excel 对齐
    5. 写入 questions 表
    """
    if not task_id:
        raise ValueError("task_id 不能为空")
    if not excel_path:
        raise ValueError("excel_path 不能为空")

    task = task_service.get_task_by_id(task_id)
    if not task:
        raise ValueError(f"任务不存在: {task_id}")

    paper_id = task.get("paper_id")
    if not paper_id:
        raise ValueError(f"任务未绑定 paper_id: {task_id}")

    try:
        task_update_status(task_id=task_id, status="running")

        dirs = storage_service.init_task_dirs(task_id)
        task_root = dirs["task_root"]
        blank_pages_dir = dirs["blank_pages"]
        solution_pages_dir = dirs["solution_pages"]
        blank_questions_dir = dirs["blank_questions"]
        solution_questions_dir = dirs["solution_questions"]

        excel_path_obj = Path(excel_path).resolve()
        if not excel_path_obj.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path_obj}")

        blank_page_files = _sorted_png_files(blank_pages_dir)
        solution_page_files = _sorted_png_files(solution_pages_dir)

        if not blank_page_files:
            raise ValueError("blank_pages 目录为空，无法执行 Step10.4")
        if not solution_page_files:
            raise ValueError("solution_pages 目录为空，无法执行 Step10.4")
        if len(blank_page_files) != len(solution_page_files):
            raise ValueError(
                f"blank_pages 与 solution_pages 页数不一致: "
                f"{len(blank_page_files)} vs {len(solution_page_files)}"
            )

        # 清空最终输出目录
        _clear_png_files(blank_questions_dir)
        _clear_png_files(solution_questions_dir)

        # 临时目录
        tmp_blank_root = task_root / "_tmp_blank_cut"
        tmp_solution_root = task_root / "_tmp_solution_cut"
        _clear_dir(tmp_blank_root)
        _clear_dir(tmp_solution_root)

        all_blank_crops: list[Path] = []
        all_solution_crops: list[Path] = []

        # -------------------------
        # 1) blank 页切题
        # -------------------------
        for page_idx, blank_page in enumerate(blank_page_files, start=1):
            page_out_dir = tmp_blank_root / f"page_{page_idx:03d}"
            page_out_dir.mkdir(parents=True, exist_ok=True)

            cut_questions_by_tencent_ocr(
                image_path=str(blank_page),
                output_dir=str(page_out_dir),
                top_pad=30,
                bottom_pad=20,
                left_pad=0,
                right_pad=0
            )

            page_crops = _collect_cut_results(
                page_out_dir,
                r"^question_\d+\.png$",
            )
            all_blank_crops.extend(page_crops)

        # -------------------------
        # 2) solution 页提取解析
        # -------------------------
        for page_idx, solution_page in enumerate(solution_page_files, start=1):
            page_out_dir = tmp_solution_root / f"page_{page_idx:03d}"
            page_out_dir.mkdir(parents=True, exist_ok=True)

            extract_analysis_by_question_blocks(
                image_path=str(solution_page),
                output_dir=str(page_out_dir),
                debug_output_name="debug_lines.png",
            )

            page_crops = _collect_cut_results(
                page_out_dir,
                r"^question_\d+_analysis\.png$",
            )

            # 3) 清洗解析图，直接覆盖原文件
            for analysis_img in page_crops:
                clean_analysis_image(
                    image_path=str(analysis_img),
                    output_path=str(analysis_img),
                    debug=False,
                    save_debug_ocr_json=False,
                    remove_blue_lines=True,
                    crop_outer_whitespace=False,
                    overlap_threshold=30,
                )

            all_solution_crops.extend(page_crops)

        if not all_blank_crops:
            raise ValueError("未从 blank_pages 中切出任何题目图片")
        if not all_solution_crops:
            raise ValueError("未从 solution_pages 中提取出任何解析图片")

        if len(all_blank_crops) != len(all_solution_crops):
            raise ValueError(
                f"题目图与解析图数量不一致: "
                f"{len(all_blank_crops)} vs {len(all_solution_crops)}"
            )

        # -------------------------
        # 4) Excel 对齐
        # -------------------------
        labels = _read_excel_question_labels(excel_path_obj)

        if len(labels) != len(all_blank_crops):
            raise ValueError(
                f"Excel 题号数量与切割结果数量不一致: "
                f"excel={len(labels)}, blank={len(all_blank_crops)}, solution={len(all_solution_crops)}"
            )

        # 先清空该 paper 旧 questions
        question_service.delete_questions_by_paper_id(paper_id)

        final_blank_paths: list[Path] = []
        final_solution_paths: list[Path] = []

        for idx, label in enumerate(labels, start=1):
            safe_label = _sanitize_filename(label)

            blank_src = all_blank_crops[idx - 1]
            solution_src = all_solution_crops[idx - 1]

            blank_dst = blank_questions_dir / f"{safe_label}.png"
            solution_dst = solution_questions_dir / f"{safe_label}.png"

            shutil.copy2(blank_src, blank_dst)
            shutil.copy2(solution_src, solution_dst)

            final_blank_paths.append(blank_dst)
            final_solution_paths.append(solution_dst)

            bbox_json = json.dumps(
                {
                    "source": "paper_cut",
                    "excel_label": label,
                    "global_index": idx,
                },
                ensure_ascii=False,
            )

            question_service.upsert_question(
                paper_id=paper_id,
                question_no=idx,
                blank_image_path=str(blank_dst),
                solution_image_path=str(solution_dst),
                bbox_json=bbox_json,
                match_status="matched",
                json_status="pending",
            )

        result = {
            "task_id": task_id,
            "paper_id": paper_id,
            "excel_path": str(excel_path_obj),
            "question_count": len(labels),
            "blank_questions_dir": str(blank_questions_dir),
            "solution_questions_dir": str(solution_questions_dir),
            "status": "success",
        }

        logger.info("generate_questions success: %s", result)
        task_update_status(
            task_id=task_id,
            status="success",
            output_path=str(solution_questions_dir),
        )
        return result

    except Exception as e:
        logger.exception("generate_questions failed: task_id=%s, error=%s", task_id, e)
        task_update_status(
            task_id=task_id,
            status="failed",
            error_message=str(e),
        )
        raise