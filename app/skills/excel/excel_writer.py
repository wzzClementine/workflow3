from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    def write_manifest_to_excel(
        self,
        manifest_path: str,
        output_path: str,
        school: str = "",
        year: str = "",
        paper_note: str = "",
    ) -> str:
        if not os.path.exists(manifest_path):
            raise FileNotFoundError(f"manifest 不存在: {manifest_path}")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        items = manifest.get("items", [])
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "试卷信息"

        self._build_template(ws)
        self._fill_header(ws, school=school, year=year, paper_note=paper_note)
        self._fill_rows(ws, items)

        wb.save(output_path)
        return output_path

    def _build_template(self, ws) -> None:
        # 列宽
        col_widths = {
            "A": 14,
            "B": 28,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 12,
            "G": 10,
            "H": 10,
            "I": 16,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 16,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 行高
        ws.row_dimensions[1].height = 24
        for r in range(2, 200):
            ws.row_dimensions[r].height = 22

        thin_gray = Side(style="thin", color="808080")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

        title_fill = PatternFill("solid", fgColor="FFFFFF")
        header_fill = PatternFill("solid", fgColor="A9D08E")
        body_fill = PatternFill("solid", fgColor="F6EFC8")

        red_font = Font(color="FF0000", bold=True)
        normal_font = Font(color="000000", bold=False)
        bold_font = Font(color="000000", bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        # 第一行提示
        ws["A1"] = "注：1、将excel文件名命名为试卷ID；2、选择题的选项直接写ABCD；3、试卷有多少题就写多少题号，不要有多余行。"
        ws["A1"].font = red_font
        ws["A1"].alignment = left
        ws.merge_cells("A1:N1")

        # 学校/年份/说明
        ws["A2"] = "学校："
        ws["A3"] = "试题时间："
        ws["A4"] = "试卷说明："

        for cell in ["A2", "A3", "A4"]:
            ws[cell].fill = header_fill
            ws[cell].font = bold_font
            ws[cell].alignment = left
            ws[cell].border = border

        for cell in ["B2", "B3", "B4", "C4"]:
            ws[cell].fill = body_fill
            ws[cell].font = normal_font
            ws[cell].alignment = left
            ws[cell].border = border

        ws["C4"] = "（说明可以为空）"

        # 表头
        headers = {
            "A5": "题号",
            "B5": "选项0",
            "C5": "选项1",
            "D5": "选项2",
            "E5": "选项3",
            "F5": "正确选项",
            "G5": "分值",
            "H5": "难度",
            "I5": "知识点",
        }

        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].fill = header_fill
            ws[cell].font = bold_font
            ws[cell].alignment = center
            ws[cell].border = border

        ws.merge_cells("I5:N5")
        ws["I5"] = "知识点（知识点可以加多项，依次向后添加）"

        # 给表体预设样式
        for row in range(6, 200):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.fill = body_fill
                cell.font = normal_font
                cell.alignment = left if col in [1, 2, 9, 10, 11, 12, 13, 14] else center
                cell.border = border

    def _fill_header(
        self,
        ws,
        school: str,
        year: str,
        paper_note: str,
    ) -> None:
        ws["B2"] = school
        ws["B3"] = year
        ws["B4"] = paper_note

    def _fill_rows(self, ws, items: list[dict[str, Any]]) -> None:
        start_row = 6

        for idx, item in enumerate(items, start=start_row):
            display_no = item.get("display_no") or item.get("parent_display_no") or ""
            answer = item.get("answer", "uncertain")
            score = item.get("score", None)
            knowledge_points = item.get("knowledge_points", []) or []
            question_type = item.get("question_type", "unknown")

            if isinstance(knowledge_points, str):
                knowledge_points = [knowledge_points] if knowledge_points.strip() else []
            elif not isinstance(knowledge_points, list):
                knowledge_points = []

            if score in ("", None):
                score = ""

            ws[f"A{idx}"] = display_no
            ws[f"B{idx}"] = self._normalize_answer_for_excel(answer, question_type)
            ws[f"F{idx}"] = 0
            ws[f"G{idx}"] = score
            ws[f"H{idx}"] = ""

            self._write_knowledge_points(ws, row=idx, knowledge_points=knowledge_points)

    def _normalize_answer_for_excel(
        self,
        answer: str,
        question_type: str,
    ) -> str:
        if answer is None:
            return "uncertain"

        text = str(answer).strip()
        if not text:
            return "uncertain"

        if question_type == "choice":
            normalized = (
                text.replace(" ", "")
                .replace("，", "")
                .replace(",", "")
                .replace("、", "")
                .upper()
            )
            return normalized

        return text

    def _write_knowledge_points(
        self,
        ws,
        row: int,
        knowledge_points: list[str],
    ) -> None:
        start_col = 9  # I列
        for i, point in enumerate(knowledge_points[:6]):
            ws.cell(row=row, column=start_col + i).value = point